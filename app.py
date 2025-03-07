from flask import Flask, request, send_file, render_template_string
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os
import time
import logging
import zipfile
from io import BytesIO
from PIL import Image as PILImage

app = Flask(__name__)

# Configuração de logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Limite de tamanho do arquivo de upload (10MB)
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10MB

# Página inicial com o formulário de upload
@app.route('/', methods=['GET'])
def index():
    return render_template_string('''
    <!DOCTYPE html>
    <html lang="pt-BR">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Upload de Arquivo</title>
        <style>
            body, html {
                margin: 0;
                padding: 0;
                height: 100%;
                font-family: Arial, sans-serif;
                background-color: #f4f4f4;
                display: flex;
                justify-content: center;
                align-items: center;
            }
            .container {
                text-align: center;
            }
            .file-input {
                background-color: black;
                padding: 30px;
                border-radius: 12px;
                box-shadow: 0 4px 15px rgba(0, 0, 0, 0.1);
                display: inline-block;
            }
            .file-input label {
                display: block;
                margin-bottom: 15px;
                font-size: 20px;
                color: wheat;
                font-weight: bold;
            }
            .file-input input[type="file"] {
                display: none;
            }
            .file-input .custom-file-upload {
                display: inline-block;
                padding: 12px 24px;
                background-color: #007bff;
                color: #fff;
                border-radius: 6px;
                cursor: pointer;
                font-size: 16px;
                transition: background-color 0.3s ease;
            }
            .file-input .custom-file-upload:hover {
                background-color: #0056b3;
            }
            .file-input .file-name {
                margin-top: 15px;
                font-size: 14px;
                color: white;
            }
        </style>
    </head>
    <body>
        <div class="container">
            <div class="file-input">
                <label for="file-upload">Escolha um arquivo para upload:</label>
                <form action="/upload" method="post" enctype="multipart/form-data" onsubmit="return validateFile()">
                    <input type="file" id="file-upload" name="file" accept=".xlsx, .xls">
                    <label for="file-upload" class="custom-file-upload">
                        Selecionar Arquivo
                    </label>
                    <div class="file-name" id="file-name">Nenhum arquivo selecionado</div>
                    <button type="submit" style="margin-top: 20px; padding: 10px 20px; background-color: #28a745; color: #fff; border: none; border-radius: 6px; cursor: pointer;">
                        Enviar e Processar
                    </button>
                </form>
            </div>
        </div>
        <script>
            document.getElementById('file-upload').addEventListener('change', function() {
                const fileName = this.files[0] ? this.files[0].name : "Nenhum arquivo selecionado";
                document.getElementById('file-name').textContent = fileName;
            });
            function validateFile() {
                const fileInput = document.getElementById('file-upload');
                if (fileInput.files.length === 0) {
                    alert("Por favor, selecione um arquivo.");
                    return false;
                }
                if (fileInput.files.length > 1) {
                    alert("Apenas um arquivo pode ser enviado por vez.");
                    return false;
                }
                return true;
            }
        </script>
    </body>
    </html>
    ''')

# Rota para processar o arquivo enviado
@app.route('/upload', methods=['POST'])
def upload_file():
    start_time = time.time()
    if 'file' not in request.files:
        logger.error("Nenhum arquivo enviado")
        return "Nenhum arquivo enviado", 400
    file = request.files['file']
    if file.filename == '':
        logger.error("Nenhum arquivo selecionado")
        return "Nenhum arquivo selecionado", 400
    if file and file.filename.endswith(('.xlsx', '.xls')):
        # Verificar o tamanho do arquivo
        file.seek(0, os.SEEK_END)
        file_length = file.tell()
        file.seek(0)
        if file_length > app.config['MAX_CONTENT_LENGTH']:
            logger.error("Arquivo excede o tamanho máximo permitido")
            return "Arquivo excede o tamanho máximo permitido (10MB)", 400

        # Processar o arquivo
        output_zip = process_file(file)

        # Enviar o arquivo processado para o usuário
        logger.info(f"Arquivo processado em {time.time() - start_time:.2f} segundos")
        return send_file(output_zip, as_attachment=True, download_name='arquivos_gerados.zip')
    else:
        logger.error("Formato de arquivo inválido")
        return "Formato de arquivo inválido. Apenas arquivos .xlsx ou .xls são permitidos.", 400

# Função para processar o arquivo Excel
def process_file(file):
    start_time = time.time()
    # Carregar o arquivo Excel
    df = pd.read_excel(file, usecols="A:H")
    df = df.dropna(how='all')  # Remover linhas completamente vazias
    dados_atualizados = [tuple(row) for row in df.itertuples(index=False, name=None)]

    # Criar um arquivo .zip em memória
    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w') as zipf:
        for or_, ta, obra, localidade, causa, tratativa, endereco, exec_obra in dados_atualizados:
            output_file = preencher_planilha(ta, obra, localidade, tratativa, endereco, exec_obra, or_, causa)
            zipf.writestr(f'{obra}.xlsx', output_file.getvalue())

    zip_buffer.seek(0)
    logger.info(f"Arquivo processado em {time.time() - start_time:.2f} segundos")
    return zip_buffer

# Função para preencher a planilha base
def preencher_planilha(ta, obra, localidade, tratativa, endereco, exec_obra, or_, causa, nome_arquivo_base='modelocroqui.xlsx'):
    start_time = time.time()
    # Carregar a planilha base
    wb = load_workbook(nome_arquivo_base)
    ws = wb.active  # Selecionar a primeira aba

    # Copiar as imagens da planilha base
    images = []
    for image in ws._images:
        img = PILImage.open(image.ref)
        img_byte_arr = BytesIO()
        img.save(img_byte_arr, format=img.format)


    # Preencher os dados nas células especificadas
    ws['C53'] = obra          # Obra
    ws['H53'] = ta            # TA
    ws['S32'] = localidade    # Localidade
    ws['B56'] = tratativa     # Tratativa
    ws['H31'] = endereco      # Endereço
    ws['L43'] = exec_obra     # Execução de Obra
    ws['C42'] = or_           # OR
    ws['C51'] = causa         # Causa

    # Adicionar as imagens de volta à planilha
    for img in images:
        ws.add_image(img)

    # Salvar a planilha atualizada em memória
    output_buffer = BytesIO()
    wb.save(output_buffer)
    output_buffer.seek(0)

    logger.info(f"Planilha preenchida em {time.time() - start_time:.2f} segundos")
    return output_buffer

# Iniciar o servidor Flask
if __name__ == "__main__":
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=True)
